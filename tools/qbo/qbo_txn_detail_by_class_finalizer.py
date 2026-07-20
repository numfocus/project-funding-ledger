#!/usr/bin/env python3
"""Finalize the Upload worksheet for transaction import.

This module runs after the transformer and dimensioner. It filters rows using
organization approval status from organization_registry.yaml, then rebuilds
the Upload worksheet in the required import layout.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


ORGANIZATION_REGISTRY_FILENAME = "organization_registry.yaml"
UPLOAD_SHEET = "Upload"

FINAL_HEADERS = [
    "PFL_Transaction_Key",
    "Organization",
    "Funding_Source",
    "Transaction_Date",
    "Transaction_Type",
    "Counter_Party",
    "Distribution_Account",
    "Memo_Description",
    "Amount",
]

REQUIRED_SOURCE_HEADERS = {
    "Transaction_Date",
    "Transaction_Type",
    "Counter_Party",
    "Memo_Description",
    "Amount",
    "Distribution_Account",
    "QBO_Project",
    "QBO_Funding_Source",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def resolve_registry_path(
    registry_path: Path | None = None,
) -> Path:
    """Resolve the organization registry path.

    Search order when no path is supplied:

    1. config/organization_registry.yaml beside this module
    2. organization_registry.yaml beside this module
    """
    if registry_path is not None:
        resolved = Path(registry_path).expanduser().resolve()
        if not resolved.exists():
            raise FileNotFoundError(
                f"Organization registry does not exist: {resolved}"
            )
        if not resolved.is_file():
            raise ValueError(
                f"Organization registry path is not a file: {resolved}"
            )
        return resolved

    module_directory = Path(__file__).resolve().parent
    candidates = (
        module_directory / "config" / ORGANIZATION_REGISTRY_FILENAME,
        module_directory / ORGANIZATION_REGISTRY_FILENAME,
    )

    for candidate in candidates:
        if candidate.is_file():
            return candidate

    searched = ", ".join(str(path) for path in candidates)
    raise FileNotFoundError(
        f"Could not find {ORGANIZATION_REGISTRY_FILENAME}. "
        f"Searched: {searched}"
    )


def load_organization_registry(
    registry_path: Path | None = None,
) -> dict[str, Any]:
    """Load the registry and derive the approved organization allowlist."""
    resolved_path = resolve_registry_path(registry_path)

    with resolved_path.open("r", encoding="utf-8") as registry_file:
        raw_registry = yaml.safe_load(registry_file) or {}

    if not isinstance(raw_registry, dict):
        raise ValueError(
            f"Organization registry must contain a YAML mapping: {resolved_path}"
        )

    if raw_registry.get("schema") != "organization_registry":
        raise ValueError(
            f"YAML file is not an organization registry: {resolved_path}"
        )

    organizations = raw_registry.get("organizations")
    if not isinstance(organizations, dict) or not organizations:
        raise ValueError(
            "Organization registry must contain a non-empty "
            "'organizations' mapping."
        )

    approved_organizations: list[str] = []
    approved_casefold: set[str] = set()

    for stable_id, record in organizations.items():
        if not isinstance(stable_id, str) or not stable_id.strip():
            raise ValueError(
                "Organization registry contains a blank stable organization ID."
            )
        if not isinstance(record, dict):
            raise ValueError(
                f"Organization {stable_id!r} must be a YAML mapping."
            )

        organization_name = text(record.get("name"))
        if not organization_name:
            raise ValueError(
                f"Organization {stable_id!r} requires a nonblank name."
            )

        approved = record.get("approved")
        if not isinstance(approved, bool):
            raise ValueError(
                f"Organization {stable_id!r} requires a boolean "
                "'approved' value."
            )

        normalized_name = organization_name.casefold()
        if normalized_name in approved_casefold:
            raise ValueError(
                "Organization registry contains duplicate approved "
                f"organization names: {organization_name}"
            )

        if approved:
            approved_organizations.append(organization_name)
            approved_casefold.add(normalized_name)

    if not approved_organizations:
        raise ValueError(
            "Organization registry contains no organizations with approved: true."
        )

    return {
        "approved_organizations": approved_organizations,
        "approved_organizations_casefold": approved_casefold,
        "registry_path": resolved_path,
    }

def header_map(ws: Worksheet) -> dict[str, int]:
    """Return a mapping of header name to one-based column number."""
    result: dict[str, int] = {}
    for cell in ws[1]:
        name = text(cell.value)
        if name:
            result[name] = cell.column
    return result


def validate_upload_sheet(ws: Worksheet) -> dict[str, int]:
    """Validate the source Upload worksheet and return its header map."""
    headers = header_map(ws)
    missing = sorted(REQUIRED_SOURCE_HEADERS - headers.keys())
    if missing:
        raise ValueError(
            "Upload worksheet is missing required column(s): " + ", ".join(missing)
        )
    return headers


def approved_organization(value: Any, approved_casefold: set[str]) -> bool:
    """Return whether a value exactly matches a configured organization."""
    return text(value).casefold() in approved_casefold


def collect_final_rows(
    ws: Worksheet,
    headers: dict[str, int],
    approved_casefold: set[str],
) -> tuple[list[list[Any]], dict[str, int]]:
    """Filter configured organization rows and build the final row set."""
    final_rows: list[list[Any]] = []
    kept = 0
    removed = 0
    skipped_blank = 0

    for row_number in range(2, ws.max_row + 1):
        organization = ws.cell(row_number, headers["QBO_Project"]).value

        if text(organization) == "":
            skipped_blank += 1
            continue

        if not approved_organization(organization, approved_casefold):
            removed += 1
            continue

        funding_source = ws.cell(
            row_number,
            headers["QBO_Funding_Source"],
        ).value

        final_rows.append(
            [
                funding_source,
                organization,
                funding_source,
                ws.cell(row_number, headers["Transaction_Date"]).value,
                ws.cell(row_number, headers["Transaction_Type"]).value,
                ws.cell(row_number, headers["Counter_Party"]).value,
                ws.cell(row_number, headers["Distribution_Account"]).value,
                ws.cell(row_number, headers["Memo_Description"]).value,
                ws.cell(row_number, headers["Amount"]).value,
            ]
        )
        kept += 1

    return final_rows, {
        "rows_kept": kept,
        "rows_removed": removed,
        "rows_skipped_blank_organization": skipped_blank,
    }


def rebuild_upload_sheet(ws: Worksheet, final_rows: list[list[Any]]) -> None:
    """Replace the worksheet contents with the final import layout."""
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    ws.append(FINAL_HEADERS)
    for row in final_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    ws.row_dimensions[1].height = 32

    widths = {
        "A": 45,
        "B": 28,
        "C": 45,
        "D": 15,
        "E": 20,
        "F": 32,
        "G": 38,
        "H": 70,
        "I": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row_number in range(2, ws.max_row + 1):
        ws.cell(row_number, 4).number_format = "mm/dd/yyyy"
        ws.cell(row_number, 9).number_format = "#,##0.00;[Red]-#,##0.00"


def finalize_upload(
    workbook_path: Path,
    registry_path: Path | None = None,
) -> dict[str, int]:
    """Finalize the Upload worksheet in place and return processing counts."""
    workbook_path = Path(workbook_path).expanduser().resolve()

    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Finalizer input must be an .xlsx workbook.")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")
    if not workbook_path.is_file():
        raise ValueError(f"Workbook path is not a file: {workbook_path}")

    registry = load_organization_registry(registry_path)
    approved_casefold = registry["approved_organizations_casefold"]

    workbook = load_workbook(workbook_path)
    try:
        if workbook.sheetnames != [UPLOAD_SHEET]:
            raise ValueError(
                'Workbook must contain exactly one worksheet named "Upload"; '
                f"found: {', '.join(workbook.sheetnames) or '(none)'}"
            )

        ws = workbook[UPLOAD_SHEET]
        headers = validate_upload_sheet(ws)
        final_rows, counts = collect_final_rows(
            ws,
            headers,
            approved_casefold,
        )
        rebuild_upload_sheet(ws, final_rows)
        workbook.save(workbook_path)
    finally:
        workbook.close()

    return counts


def main(argv: list[str]) -> int:
    if len(argv) not in (1, 2):
        print(
            "Usage: python3 qbo_txn_detail_by_class_finalizer.py "
            "<workbook.xlsx> [organization_registry.yaml]",
            file=sys.stderr,
        )
        return 2

    workbook_path = Path(argv[0])
    registry_path = Path(argv[1]) if len(argv) == 2 else None

    try:
        counts = finalize_upload(workbook_path, registry_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Finalized: {workbook_path.expanduser().resolve()}")
    print(f"Rows kept: {counts['rows_kept']}")
    print(f"Rows removed: {counts['rows_removed']}")
    print(
        "Rows skipped for blank Organization: "
        f"{counts['rows_skipped_blank_organization']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
