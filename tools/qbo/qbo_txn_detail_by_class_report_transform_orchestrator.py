#!/usr/bin/env python3
"""Orchestrate the QBO transaction-detail-by-class transformation pipeline.

Command-line parameters use Key=Value syntax and may be supplied in any order.

Required parameters:

    InputFile=/path/to/source.xlsx
    OutputDirectory=/path/to/output
    OutputFilename=output_file_name
    YAMLFileDirectory=/path/to/yaml/files

The YAML directory must contain:

    organization_registry.yaml

The transformer and dimensioner receive the YAML directory and are responsible
for validating their own additional YAML dependencies.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

import qbo_txn_detail_by_class_dimensioner as dimensioner
from qbo_txn_detail_by_class_finalizer import finalize_upload
from qbo_txn_detail_by_class_transformer import transform_report


ORGANIZATION_REGISTRY_FILENAME = "organization_registry.yaml"

REQUIRED_PARAMETERS = {
    "InputFile",
    "OutputDirectory",
    "OutputFilename",
    "YAMLFileDirectory",
}


@dataclass(frozen=True)
class PipelinePaths:
    input_path: Path
    output_directory: Path
    output_path: Path
    yaml_directory: Path
    organization_registry_path: Path


def parse_key_value_argument(argument: str) -> tuple[str, str]:
    """Parse one Key=Value argument."""
    if "=" not in argument:
        raise argparse.ArgumentTypeError(
            f"Invalid parameter {argument!r}. Expected Key=Value syntax."
        )

    key, value = argument.split("=", 1)
    key = key.strip()
    value = value.strip()

    if not key:
        raise argparse.ArgumentTypeError(
            f"Invalid parameter {argument!r}. Parameter name cannot be blank."
        )
    if not value:
        raise argparse.ArgumentTypeError(
            f"Invalid parameter {argument!r}. Parameter value cannot be blank."
        )

    return key, value


def parse_args(argv: list[str]) -> dict[str, str]:
    parser = argparse.ArgumentParser(
        description=(
            "Run the complete QBO transaction migration pipeline using "
            "Key=Value parameters."
        ),
        epilog=(
            'Example: python3 qbo_txn_detail_by_class_report_transform_orchestrator.py '
            '"InputFile=/path/source.xlsx" '
            '"OutputDirectory=/path/out" '
            '"OutputFilename=Upload_Transactions" '
            '"YAMLFileDirectory=/path/yaml"'
        ),
    )
    parser.add_argument(
        "parameters",
        nargs="+",
        metavar="Key=Value",
        help="Named pipeline parameter in Key=Value format.",
    )

    namespace = parser.parse_args(argv)
    parsed: dict[str, str] = {}

    for raw_argument in namespace.parameters:
        try:
            key, value = parse_key_value_argument(raw_argument)
        except argparse.ArgumentTypeError as exc:
            parser.error(str(exc))

        if key not in REQUIRED_PARAMETERS:
            parser.error(
                f"Unknown parameter {key!r}. Expected: "
                + ", ".join(sorted(REQUIRED_PARAMETERS))
            )

        if key in parsed:
            parser.error(f"Parameter {key!r} was supplied more than once.")

        parsed[key] = value

    missing = sorted(REQUIRED_PARAMETERS - parsed.keys())
    if missing:
        parser.error("Missing required parameter(s): " + ", ".join(missing))

    return parsed


def validate_source_workbook(input_path: Path) -> None:
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Input file must be an .xlsx or .xlsm workbook.")
    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_path}")
    if not input_path.is_file():
        raise ValueError(f"Input path is not a file: {input_path}")

    workbook = load_workbook(input_path, read_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError(
                "Input workbook must contain exactly one worksheet; "
                f"found {len(workbook.sheetnames)}."
            )
    finally:
        workbook.close()


def normalize_output_name(name: str) -> str:
    name = name.strip()
    if not name:
        raise ValueError("Output filename cannot be blank.")
    if Path(name).name != name:
        raise ValueError("Output filename must not contain directories.")
    if Path(name).suffix and Path(name).suffix.lower() != ".xlsx":
        raise ValueError("Output filename must use the .xlsx extension.")
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name


def validate_yaml_file(path: Path, description: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{description} does not exist: {path}")
    if not path.is_file():
        raise ValueError(f"{description} is not a file: {path}")


def validate_parameters(parameters: dict[str, str]) -> PipelinePaths:
    input_path = Path(parameters["InputFile"]).expanduser().resolve()
    output_directory = Path(parameters["OutputDirectory"]).expanduser().resolve()
    yaml_directory = Path(parameters["YAMLFileDirectory"]).expanduser().resolve()

    validate_source_workbook(input_path)

    if output_directory.exists() and not output_directory.is_dir():
        raise ValueError(
            f"Output directory path is not a directory: {output_directory}"
        )
    output_directory.mkdir(parents=True, exist_ok=True)

    if not yaml_directory.exists():
        raise FileNotFoundError(f"YAML directory does not exist: {yaml_directory}")
    if not yaml_directory.is_dir():
        raise ValueError(f"YAML path is not a directory: {yaml_directory}")

    organization_registry_path = (
        yaml_directory / ORGANIZATION_REGISTRY_FILENAME
    )
    validate_yaml_file(
        organization_registry_path,
        "Organization registry YAML file",
    )

    output_path = output_directory / normalize_output_name(
        parameters["OutputFilename"]
    )

    if output_path.resolve() == input_path.resolve():
        raise ValueError("Output file cannot overwrite the source workbook.")

    return PipelinePaths(
        input_path=input_path,
        output_directory=output_directory,
        output_path=output_path,
        yaml_directory=yaml_directory,
        organization_registry_path=organization_registry_path,
    )


def run_pipeline(
    paths: PipelinePaths,
) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    print("Step 1 complete: Parameters and source workbook validated.")
    print(f"YAML directory: {paths.yaml_directory}")

    print("Step 2: Transforming report...")
    transform_counts = transform_report(
        paths.input_path,
        paths.output_path,
        paths.yaml_directory,
    )

    print("Step 3: Populating QBO dimensions...")
    dimension_counts = dimensioner.dimension_upload(
        paths.output_path,
        paths.yaml_directory,
    )

    print("Step 4: Finalizing Upload worksheet...")
    finalizer_counts = finalize_upload(
        paths.output_path,
        paths.organization_registry_path,
    )

    return transform_counts, dimension_counts, finalizer_counts


def main(argv: list[str]) -> int:
    try:
        parameters = parse_args(argv)
        paths = validate_parameters(parameters)
        transform_counts, dimension_counts, finalizer_counts = run_pipeline(paths)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print("\nPipeline completed successfully.")
    print(f"Saved: {paths.output_path}")
    print(transform_counts)
    print(dimension_counts)
    print(finalizer_counts)

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
