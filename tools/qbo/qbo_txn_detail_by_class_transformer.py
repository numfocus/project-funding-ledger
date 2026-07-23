#!/usr/bin/env python3
"""Report transformation logic for QBO transaction-detail-by-class exports.

This module is called by the pipeline orchestrator. It does not parse command-line
arguments. It reads the validated source workbook, loads shared configuration
from organization_registry.yaml, creates a new workbook containing one worksheet
named "Upload", and saves it to the output path supplied by the orchestrator.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ORGANIZATION_REGISTRY_FILENAME = "organization_registry.yaml"

SOURCE_HEADERS = [
    "Class_Blocks",
    "Transaction_Date",
    "Transaction_Type",
    "Num",
    "Counter_Party",
    "Product_Service",
    "Memo_Description",
    "Quantity",
    "Price",
    "Amount",
    "Account_Type",
    "Class_Full_Name",
    "Line_Currency",
    "Account_Full_Name",
    "Distribution_Account",
]

OUTPUT_HEADERS = SOURCE_HEADERS + [
    "Row_Identifier",
    "Derived_Class_1",
    "Backup_Class_2",
    "Data_Row_1",
    "Special_Program_Prefix",
    "Derived_Class_2",
    "Remove_Deleted",
    "SDG_Round_1",
    "SDG_Round",
    "Derived_Class",
    "QBO_Project",
    "QBO_Class",
    "QBO_Location",
    "QBO_Funding_Source",
    "QBO_Service_Agreement",
    "QBO_Program_Initiative",
    "QBO_Account",
    "Deleted_Class",
    "Review_Flag",
]

EXPECTED_DOWNLOAD_HEADERS = {
    2: "Transaction date",
    3: "Transaction type",
    12: "Class full name",
    15: "Distribution account",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()



def load_special_program_rules(
    yaml_directory: Path,
) -> dict[str, Any]:
    """Load and validate cross-organization prefix rules from the registry."""
    yaml_path = (
        Path(yaml_directory).expanduser().resolve()
        / ORGANIZATION_REGISTRY_FILENAME
    )

    if not yaml_path.exists():
        raise FileNotFoundError(
            f"Required organization registry not found: {yaml_path}"
        )
    if not yaml_path.is_file():
        raise ValueError(
            f"Organization registry path is not a file: {yaml_path}"
        )

    try:
        with yaml_path.open("r", encoding="utf-8") as yaml_file:
            config = yaml.safe_load(yaml_file) or {}
    except yaml.YAMLError as exc:
        raise ValueError(f"Invalid YAML in {yaml_path}: {exc}") from exc

    if not isinstance(config, dict):
        raise ValueError(
            f"Organization registry root must be a mapping: {yaml_path}"
        )

    if config.get("schema") != "organization_registry":
        raise ValueError(
            f"YAML file is not an organization registry: {yaml_path}"
        )

    organizations = config.get("organizations")
    if not isinstance(organizations, dict) or not organizations:
        raise ValueError(
            f"'organizations' must be a nonempty mapping in {yaml_path}"
        )

    options = config.get("options", {})
    if options is None:
        options = {}
    if not isinstance(options, dict):
        raise ValueError(
            f"'options' must be a mapping in {yaml_path}"
        )

    ignore_case = options.get("ignore_case", True)
    if not isinstance(ignore_case, bool):
        raise ValueError(
            f"'options.ignore_case' must be true or false in {yaml_path}"
        )

    rules = config.get("special_programs", [])
    if rules is None:
        rules = []
    if not isinstance(rules, list):
        raise ValueError(
            f"'special_programs' must be a list in {yaml_path}"
        )

    validated_rules: list[dict[str, Any]] = []

    for index, rule in enumerate(rules, start=1):
        if not isinstance(rule, dict):
            raise ValueError(
                f"special_programs entry {index} must be a mapping "
                f"in {yaml_path}"
            )

        prefix = rule.get("prefix")
        match_values = rule.get("match")

        if not isinstance(prefix, str) or not prefix.strip():
            raise ValueError(
                f"special_programs entry {index} requires "
                "a nonblank 'prefix'"
            )

        if not isinstance(match_values, list) or not match_values:
            raise ValueError(
                f"special_programs entry {index} requires "
                "a nonempty 'match' list"
            )

        normalized_matches: list[str] = []

        for match_index, match_value in enumerate(
            match_values,
            start=1,
        ):
            if not isinstance(match_value, str) or not match_value.strip():
                raise ValueError(
                    "special_programs entry "
                    f"{index}, match item {match_index} "
                    "must be nonblank text"
                )

            normalized_match = match_value.strip()

            if ignore_case:
                normalized_match = normalized_match.casefold()

            normalized_matches.append(normalized_match)

        validated_rules.append(
            {
                "prefix": prefix.strip(),
                "match": normalized_matches,
            }
        )

    return {
        "ignore_case": ignore_case,
        "rules": validated_rules,
    }


def find_download_header_row(ws: Worksheet) -> int:
    """Locate the QBO column-heading row instead of assuming a fixed row."""
    for row in range(1, min(ws.max_row, 50) + 1):
        matches = 0
        for column, expected in EXPECTED_DOWNLOAD_HEADERS.items():
            if text(ws.cell(row, column).value).casefold() == expected.casefold():
                matches += 1
        if matches == len(EXPECTED_DOWNLOAD_HEADERS):
            return row

    raise ValueError(
        'Could not find the QBO header row. Expected columns including '
        '"Transaction date", "Class full name", and "Distribution account".'
    )


def special_program_prefix(
    *values: Any,
    rules: dict[str, Any],
) -> str:
    """Return the first configured prefix whose match text appears in the values."""
    combined = " | ".join(text(value) for value in values)

    if rules["ignore_case"]:
        combined = combined.casefold()

    for rule in rules["rules"]:
        if any(match_text in combined for match_text in rule["match"]):
            return str(rule["prefix"])

    return ""


def strip_deleted_marker(value: Any) -> str:
    """Discard '(deleted)' and all text following it."""
    value_text = text(value)
    match = re.search(r"\s*\(deleted\)", value_text, flags=re.IGNORECASE)
    if match:
        value_text = value_text[: match.start()]
    return " ".join(value_text.split())


def extract_sdg_round(value: Any) -> str:
    """Return text beginning with SDG, such as 'SDG R1 2025'."""
    value_text = text(value)
    match = re.search(r"\bSDG\b.*$", value_text, flags=re.IGNORECASE)
    return " ".join(match.group(0).split()) if match else ""


def is_transaction_row(transaction_date: Any, carried_class: str) -> bool:
    """Exclude QBO headings, subtotal rows, footer rows, and blank rows."""
    date_text = text(transaction_date)
    if not date_text:
        return False
    if "total" in date_text.casefold():
        return False
    if "total" in carried_class.casefold():
        return False
    return True


def prepare_upload_sheet(
    ws_source: Worksheet,
    ws_upload: Worksheet,
    special_program_rules: list[dict[str, Any]],
    starting_identifier: int = 100002,
) -> dict[str, int]:
    header_row = find_download_header_row(ws_source)
    current_class = ""
    source_rows = 0
    transaction_rows = 0
    excluded_rows = 0

    ws_upload.append(OUTPUT_HEADERS)

    for source_row in range(header_row + 1, ws_source.max_row + 1):
        source_rows += 1
        source_values = [ws_source.cell(source_row, col).value for col in range(1, 16)]

        class_block = text(source_values[0])
        if class_block:
            current_class = class_block

        transaction_date = source_values[1]
        if not is_transaction_row(transaction_date, current_class):
            excluded_rows += 1
            continue

        modifier = special_program_prefix(
            current_class,
            source_values[11],
            source_values[6],
            rules=special_program_rules,
        )
        derived_class_2 = f"{modifier}{current_class}"
        cleaned_class = strip_deleted_marker(derived_class_2)
        sdg_round = extract_sdg_round(cleaned_class)

        output_values = source_values + [
            starting_identifier + (source_row - (header_row + 1)),
            current_class,
            "",
            "KEEP",
            modifier,
            derived_class_2,
            cleaned_class,
            sdg_round,
            sdg_round,
            cleaned_class,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        ws_upload.append(output_values)
        transaction_rows += 1

    for cell in ws_upload[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_upload.freeze_panes = "A2"
    ws_upload.auto_filter.ref = f"A1:AH{ws_upload.max_row}"
    ws_upload.row_dimensions[1].height = 32

    widths = {
        "A": 30, "B": 15, "C": 20, "D": 18, "E": 30, "F": 24,
        "G": 60, "H": 12, "I": 12, "J": 14, "K": 20, "L": 55,
        "M": 14, "N": 55, "O": 36, "P": 16, "Q": 38, "R": 30,
        "S": 14, "T": 22, "U": 42, "V": 42, "W": 20, "X": 20,
        "Y": 42, "Z": 30, "AA": 22, "AB": 24, "AC": 45,
        "AD": 35, "AE": 35, "AF": 30, "AG": 30, "AH": 40,
    }
    for column, width in widths.items():
        ws_upload.column_dimensions[column].width = width

    for row in range(2, ws_upload.max_row + 1):
        ws_upload.cell(row, 2).number_format = "mm/dd/yyyy"
        ws_upload.cell(row, 9).number_format = "#,##0.00"
        ws_upload.cell(row, 10).number_format = "#,##0.00;[Red]-#,##0.00"

    return {
        "source_rows_examined": source_rows,
        "transaction_rows_written": transaction_rows,
        "non_transaction_rows_excluded": excluded_rows,
    }


def transform_report(
    input_path: Path,
    output_path: Path,
    yaml_directory: Path,
) -> dict[str, int]:
    """Create the initial Upload workbook from a validated QBO report."""
    special_program_rules = load_special_program_rules(yaml_directory)

    source_workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        source_sheet = source_workbook[source_workbook.sheetnames[0]]

        output_workbook = Workbook()
        try:
            upload_sheet = output_workbook.active
            upload_sheet.title = "Upload"

            counts = prepare_upload_sheet(
                source_sheet,
                upload_sheet,
                special_program_rules,
            )
            output_workbook.save(output_path)
        finally:
            output_workbook.close()
    finally:
        source_workbook.close()

    return counts
