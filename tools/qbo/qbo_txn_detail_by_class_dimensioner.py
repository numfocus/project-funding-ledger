#!/usr/bin/env python3
"""Populate QBO dimension fields on the Upload worksheet.

This module is called after qbo_txn_detail_by_class_transformer.py creates the
output workbook. It modifies that workbook in place and populates:

    QBO_Project
    QBO_Class
    QBO_Location
    QBO_Funding_Source
    QBO_Service_Agreement
    QBO_Program_Initiative
    QBO_Account

Rows that cannot be mapped are marked in Review_Flag.
"""
from __future__ import annotations

import sys
from datetime import datetime, date
from functools import lru_cache
from pathlib import Path
from typing import Any, Optional

import yaml

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

COL = {
    "Transaction_Date": "B",
    "Counter_Party": "E",
    "Memo_Description": "G",
    "Class_Full_Name": "L",
    "Distribution_Account": "O",
    "Backup_Class_2": "R",
    "Data_Row_1": "S",
    "SDG_Round": "X",
    "Derived_Class": "Y",
    "QBO_Project": "Z",
    "QBO_Class": "AA",
    "QBO_Location": "AB",
    "QBO_Funding_Source": "AC",
    "QBO_Service_Agreement": "AD",
    "QBO_Program_Initiative": "AE",
    "QBO_Account": "AF",
    "Deleted_Class": "AG",
    "Review_Flag": "AH",
}

DEFAULT_CLASS = "Program Services"
DEFAULT_LOCATION = "Fiscal Sponsorship"
DEFAULT_FUNDING = "General - Unrestricted"
EVENT_MAPPING_FILENAME = "qbo_event_mapping.yaml"
PROGRAM_MAPPING_FILENAME = "qbo_program_mapping.yaml"
SDG_MAPPING_FILENAME = "qbo_sdg_mapping.yaml"
ADMIN_FUNDRAISING_MAPPING_FILENAME = "qbo_admin_fundraising_mapping.yaml"
ORGANIZATION_REGISTRY_FILENAME = "organization_registry.yaml"

_yaml_directory: Optional[Path] = None


def configure_yaml_directory(yaml_directory: Path) -> None:
    """Configure the directory containing all dimensioner YAML files."""
    global _yaml_directory

    resolved = Path(yaml_directory).expanduser().resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"YAML directory not found: {resolved}")
    if not resolved.is_dir():
        raise ValueError(f"YAML path is not a directory: {resolved}")

    _yaml_directory = resolved


def yaml_mapping_path(filename: str) -> Path:
    """Return a required mapping file from the configured YAML directory."""
    if _yaml_directory is None:
        raise RuntimeError("The dimensioner YAML directory has not been configured.")

    mapping_path = _yaml_directory / filename
    if not mapping_path.exists():
        raise FileNotFoundError(f"Required YAML mapping file not found: {mapping_path}")
    if not mapping_path.is_file():
        raise ValueError(f"YAML mapping path is not a file: {mapping_path}")
    return mapping_path


def clear_mapping_caches() -> None:
    """Clear cached mappings after changing the configured YAML directory."""
    load_sdg_rule.cache_clear()
    load_program_rules.cache_clear()
    load_admin_fundraising_rules.cache_clear()
    load_event_rules.cache_clear()
    load_organization_registry.cache_clear()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return text(value).casefold()


def starts(value: str, prefix: str) -> bool:
    return norm(value).startswith(prefix.casefold())


def contains(value: str, needle: str) -> bool:
    return needle.casefold() in norm(value)


def clean_deleted_marker(value: str) -> str:
    """Remove legacy QBO '(deleted)' marker from new dimension fields."""
    cleaned = text(value).replace("(deleted)", "")
    return " ".join(cleaned.split())


def is_number(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    try:
        float(str(value).strip())
        return True
    except Exception:
        return False


def as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if is_blank(value):
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    return None


def on_or_after(value: Any, year: int, month: int, day: int) -> bool:
    parsed = as_date(value)
    return parsed is not None and parsed >= date(year, month, day)


def fiscal_year_by_cutoff(value: Any, cutoffs: list[tuple[date, int]], fallback: int) -> int:
    parsed = as_date(value)
    if parsed is None:
        return fallback
    for cutoff, year in sorted(cutoffs, reverse=True):
        if parsed >= cutoff:
            return year
    return fallback


def year_from_text(value: str, fallback: Optional[int] = None) -> Optional[int]:
    for year in (2027, 2026, 2025, 2024, 2023):
        if str(year) in value:
            return year
    return fallback


def set_qbo(ws: Worksheet, row: int, project: str, qbo_class: str, location: str,
            funding_source: str, service_agreement: str = "", program_initiative: str = "") -> None:
    ws[f"{COL['QBO_Project']}{row}"] = clean_deleted_marker(project)
    ws[f"{COL['QBO_Class']}{row}"] = qbo_class
    ws[f"{COL['QBO_Location']}{row}"] = location
    ws[f"{COL['QBO_Funding_Source']}{row}"] = clean_deleted_marker(funding_source)
    ws[f"{COL['QBO_Service_Agreement']}{row}"] = clean_deleted_marker(service_agreement)
    ws[f"{COL['QBO_Program_Initiative']}{row}"] = clean_deleted_marker(program_initiative)


def append_memo(ws: Worksheet, row: int, addition: str) -> None:
    cell = ws[f"{COL['Memo_Description']}{row}"]
    current = text(cell.value)
    if not addition:
        return
    if addition in current:
        return
    cell.value = addition if current == "" else f"{current} | {addition}"


def preserve_legacy_detail(ws: Worksheet, row: int) -> None:
    class_full_name = text(ws[f"{COL['Class_Full_Name']}{row}"].value)
    backup_class = text(ws[f"{COL['Backup_Class_2']}{row}"].value)
    derived_class = text(ws[f"{COL['Derived_Class']}{row}"].value)
    distribution_account = text(ws[f"{COL['Distribution_Account']}{row}"].value)
    save_class = class_full_name or backup_class or derived_class
    append_memo(ws, row, f"Legacy Class: {save_class} | Legacy Distribution Account: {distribution_account}")


def _sdg_mapping_path() -> Path:
    """Return the configured SDG mapping file."""
    return yaml_mapping_path(SDG_MAPPING_FILENAME)


@lru_cache(maxsize=1)
def load_sdg_rule() -> dict[str, Any]:
    """Load and validate the SDG business rule from qbo_sdg_mapping.yaml."""
    mapping_path = _sdg_mapping_path()
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Required SDG mapping file not found: {mapping_path}"
        )

    with mapping_path.open("r", encoding="utf-8") as stream:
        configuration = yaml.safe_load(stream)

    if not isinstance(configuration, dict):
        raise ValueError(f"SDG mapping must contain a YAML object: {mapping_path}")

    rule = configuration.get("sdg")
    if not isinstance(rule, dict):
        raise ValueError(f"SDG mapping must contain an 'sdg' object: {mapping_path}")

    if not isinstance(rule.get("detection"), dict):
        raise ValueError("SDG mapping is missing its detection block.")
    if not isinstance(rule.get("year"), dict):
        raise ValueError("SDG mapping is missing its year block.")
    if not isinstance(rule.get("output"), dict):
        raise ValueError("SDG mapping is missing its output block.")

    conditions = rule["detection"].get("conditions")
    if not isinstance(conditions, list) or not conditions:
        raise ValueError("SDG detection must contain a non-empty conditions list.")

    return rule


def sdg_condition_matches(
    condition: dict[str, Any],
    derived_class: str,
    distribution_account: str,
) -> bool:
    """Evaluate one configured SDG detection condition."""
    field_values = {
        "derived_class": derived_class,
        "distribution_account": distribution_account,
    }
    field_name = text(condition.get("field"))
    if field_name not in field_values:
        raise ValueError(f"Unsupported SDG detection field: {field_name}")

    match_type = text(condition.get("match"))
    expected = text(condition.get("value"))
    actual = field_values[field_name]

    if match_type == "exact":
        return actual == expected
    if match_type == "contains":
        return contains(actual, expected)

    raise ValueError(f"Unsupported SDG detection match type: {match_type}")


def resolve_sdg_year(
    year_rule: dict[str, Any],
    derived_class: str,
    transaction_date: Any,
) -> int:
    """Resolve the SDG year entirely from the configured YAML rule."""
    mode = text(year_rule.get("mode"))
    if mode != "source_then_ranges":
        raise ValueError(f"Unsupported SDG year mode: {mode}")

    source = year_rule.get("source", {})
    if not isinstance(source, dict):
        raise ValueError("SDG year source must be a YAML object.")

    source_field = text(source.get("field"))
    if source_field != "derived_class":
        raise ValueError(f"Unsupported SDG year source field: {source_field}")

    source_years = source.get("years", [])
    if not isinstance(source_years, list):
        raise ValueError("SDG source years must be a list.")
    for configured_year in source_years:
        year = int(configured_year)
        if str(year) in derived_class:
            return year

    parsed_date = as_date(transaction_date)
    if parsed_date is not None:
        ranges = year_rule.get("ranges", [])
        if not isinstance(ranges, list):
            raise ValueError("SDG year ranges must be a list.")
        for date_range in ranges:
            if not isinstance(date_range, dict):
                raise ValueError("Each SDG year range must be a YAML object.")
            range_start = as_date(date_range.get("start"))
            range_end = as_date(date_range.get("end"))
            if range_start is not None and parsed_date < range_start:
                continue
            if range_end is not None and parsed_date > range_end:
                continue
            return int(date_range["year"])

    return int(year_rule["fallback"])


def resolve_sdg_mapping(
    ws: Worksheet,
    row: int,
    derived_class: str,
    distribution_account: str,
    transaction_date: Any,
) -> Optional[dict[str, str]]:
    """Return configured SDG output values when the row matches the YAML rule."""
    rule = load_sdg_rule()
    detection = rule["detection"]
    conditions = detection["conditions"]
    results = [
        sdg_condition_matches(condition, derived_class, distribution_account)
        for condition in conditions
    ]

    logic = text(detection.get("logic"))
    if logic == "any":
        matched = any(results)
    elif logic == "all":
        matched = all(results)
    else:
        raise ValueError(f"Unsupported SDG detection logic: {logic}")

    if not matched:
        return None

    legacy_round = rule.get("legacy_round", {})
    if isinstance(legacy_round, dict) and legacy_round.get("append_when_present"):
        source_column = text(legacy_round.get("source_column"))
        if source_column not in COL:
            raise ValueError(f"Unknown SDG legacy-round source column: {source_column}")
        round_value = text(ws[f"{COL[source_column]}{row}"].value)
        if round_value:
            memo_template = text(legacy_round.get("memo_template"))
            append_memo(ws, row, memo_template.format(value=round_value))

    year = resolve_sdg_year(rule["year"], derived_class, transaction_date)
    output = rule["output"]
    return {
        "initiative_marker": text(output.get("initiative_marker")),
        "funding_base": text(output.get("funding_base_template")).format(year=year),
        "program_initiative": text(
            output.get("program_initiative_template")
        ).format(year=year),
    }


def handle_standalone_sdg_mapping(
    ws: Worksheet,
    row: int,
    derived_class: str,
    distribution_account: str,
    transaction_date: Any,
) -> bool:
    """Apply an SDG rule that does not belong to a configured project row."""
    rule = load_sdg_rule()
    standalone = rule.get("standalone")
    if not isinstance(standalone, dict):
        return False

    match = standalone.get("match")
    if not isinstance(match, dict):
        raise ValueError("SDG standalone mapping is missing its match block.")

    match_type = text(match.get("type"))
    if match_type == "exact":
        matched = derived_class == text(match.get("value"))
    elif match_type == "one_of":
        values = match.get("values", [])
        matched = isinstance(values, list) and derived_class in values
    else:
        raise ValueError(f"Unsupported standalone SDG match type: {match_type}")

    if not matched:
        return False

    sdg_mapping = resolve_sdg_mapping(
        ws,
        row,
        derived_class,
        distribution_account,
        transaction_date,
    )
    if sdg_mapping is None:
        raise ValueError(
            f"The configured SDG detection rule did not match standalone class: "
            f"{derived_class}"
        )

    project = text(standalone.get("project"))
    if project == "":
        raise ValueError("SDG standalone mapping requires a project.")

    def resolve_default(value: Any) -> str:
        rendered = text(value)
        rendered = rendered.replace("__DEFAULT_CLASS__", DEFAULT_CLASS)
        rendered = rendered.replace("__DEFAULT_LOCATION__", DEFAULT_LOCATION)
        rendered = rendered.replace("__DEFAULT_FUNDING__", DEFAULT_FUNDING)
        return rendered

    set_qbo(
        ws,
        row,
        project,
        resolve_default(standalone.get("qbo_class")) or DEFAULT_CLASS,
        resolve_default(standalone.get("location")) or DEFAULT_LOCATION,
        organization_funding_source(project, sdg_mapping["funding_base"]),
        resolve_default(standalone.get("service_agreement")),
        sdg_mapping["program_initiative"],
    )
    return True


def _program_mapping_path() -> Path:
    """Return the configured program mapping file."""
    return yaml_mapping_path(PROGRAM_MAPPING_FILENAME)


@lru_cache(maxsize=1)
def load_program_rules() -> tuple[dict[str, Any], ...]:
    """Load and validate program business rules from qbo_program_mapping.yaml."""
    mapping_path = _program_mapping_path()
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Required program mapping file not found: {mapping_path}"
        )

    with mapping_path.open("r", encoding="utf-8") as stream:
        configuration = yaml.safe_load(stream)

    if not isinstance(configuration, dict):
        raise ValueError(
            f"Program mapping must contain a YAML object: {mapping_path}"
        )

    programs = configuration.get("programs")
    if not isinstance(programs, list) or not programs:
        raise ValueError(
            f"Program mapping must contain a non-empty 'programs' list: {mapping_path}"
        )

    for index, rule in enumerate(programs, start=1):
        if not isinstance(rule, dict):
            raise ValueError(f"Program rule {index} must be a YAML object.")
        if not isinstance(rule.get("match"), dict):
            raise ValueError(f"Program rule {index} is missing its match block.")
        if not isinstance(rule.get("project"), dict):
            raise ValueError(f"Program rule {index} is missing its project block.")
        if not isinstance(rule.get("year"), dict):
            raise ValueError(f"Program rule {index} is missing its year block.")
        if not isinstance(rule.get("output"), dict):
            raise ValueError(f"Program rule {index} is missing its output block.")

    return tuple(programs)


def program_rule_matches(rule: dict[str, Any], derived_class: str) -> bool:
    """Return True when Derived_Class satisfies a configured program rule."""
    match = rule["match"]
    match_type = text(match.get("type"))

    if match_type == "exact":
        return derived_class == text(match.get("value"))

    if match_type == "prefix_any":
        values = match.get("values", [])
        return isinstance(values, list) and any(
            starts(derived_class, text(value)) for value in values
        )

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(
        f"Unsupported program match type in rule {rule_id}: {match_type}"
    )


def resolve_program_project(rule: dict[str, Any], derived_class: str) -> str:
    """Resolve the QBO project according to the rule's project policy."""
    project_rule = rule["project"]
    mode = text(project_rule.get("mode"))

    if mode == "fixed":
        project = text(project_rule.get("value"))
        if project:
            return project

    elif mode == "contains_candidate":
        candidates = project_rule.get("candidates", [])
        if isinstance(candidates, list):
            for candidate in candidates:
                candidate_text = text(candidate)
                if candidate_text and contains(derived_class, candidate_text):
                    return candidate_text
        fallback = text(project_rule.get("fallback"))
        if fallback:
            return fallback

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(
        f"Unable to resolve project for program rule {rule_id}; mode={mode!r}."
    )


def _year_from_ranges(year_rule: dict[str, Any], transaction_date: Any) -> Optional[int]:
    parsed_date = as_date(transaction_date)
    if parsed_date is None:
        return None

    ranges = year_rule.get("ranges", [])
    if not isinstance(ranges, list):
        return None

    for date_range in ranges:
        if not isinstance(date_range, dict):
            continue
        range_start = as_date(date_range.get("start"))
        range_end = as_date(date_range.get("end"))
        if range_start is not None and parsed_date < range_start:
            continue
        if range_end is not None and parsed_date > range_end:
            continue
        return int(date_range["year"])

    return None


def resolve_program_year(
    rule: dict[str, Any],
    derived_class: str,
    transaction_date: Any,
) -> Optional[int]:
    """Resolve the year according to the configured program year policy."""
    year_rule = rule["year"]
    mode = text(year_rule.get("mode"))

    if mode == "none":
        return None

    if mode == "fixed":
        return int(year_rule["value"])

    if mode == "fixed_or_source_then_ranges":
        source_years = year_rule.get("source_years", [])
        if isinstance(source_years, list):
            for source_year in source_years:
                year = int(source_year)
                if str(year) in derived_class:
                    return year
        range_year = _year_from_ranges(year_rule, transaction_date)
        if range_year is not None:
            return range_year
        return int(year_rule["fallback"])

    if mode == "ranges":
        range_year = _year_from_ranges(year_rule, transaction_date)
        if range_year is not None:
            return range_year
        return int(year_rule["fallback"])

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(
        f"Unsupported program year mode in rule {rule_id}: {mode}"
    )


def resolve_program_variant(rule: dict[str, Any], derived_class: str) -> dict[str, str]:
    """Return values from the first matching source-prefix variant."""
    variants = rule.get("variants", [])
    if variants is None:
        return {}
    if not isinstance(variants, list):
        rule_id = text(rule.get("id")) or "(unnamed)"
        raise ValueError(f"Program variants must be a list in rule {rule_id}.")

    for variant in variants:
        if not isinstance(variant, dict):
            continue
        source_prefix = text(variant.get("source_prefix"))
        if source_prefix and starts(derived_class, source_prefix):
            values = variant.get("values", {})
            if not isinstance(values, dict):
                rule_id = text(rule.get("id")) or "(unnamed)"
                raise ValueError(
                    f"Program variant values must be an object in rule {rule_id}."
                )
            return {str(key): text(value) for key, value in values.items()}

    return {}


def render_program_template(template: Any, values: dict[str, Any]) -> str:
    """Render a program output template and resolve configured default tokens."""
    rendered = text(template)
    rendered = rendered.replace("__DEFAULT_CLASS__", DEFAULT_CLASS)
    rendered = rendered.replace("__DEFAULT_LOCATION__", DEFAULT_LOCATION)
    rendered = rendered.replace("__DEFAULT_FUNDING__", DEFAULT_FUNDING)
    return rendered.format(**values)


def handle_program_mapping(
    ws: Worksheet,
    row: int,
    derived_class: str,
    transaction_date: Any,
) -> bool:
    """Apply the first matching rule from qbo_program_mapping.yaml."""
    for rule in load_program_rules():
        if not program_rule_matches(rule, derived_class):
            continue

        project = resolve_program_project(rule, derived_class)
        year = resolve_program_year(rule, derived_class, transaction_date)
        variant_values = resolve_program_variant(rule, derived_class)

        template_values: dict[str, Any] = {
            "project": project,
            "year": "" if year is None else year,
            "funding_program": "",
            "initiative_prefix": "",
        }
        template_values.update(variant_values)

        output = rule["output"]
        set_qbo(
            ws,
            row,
            render_program_template(
                output.get("qbo_project_template"), template_values
            ),
            render_program_template(output.get("qbo_class"), template_values),
            render_program_template(output.get("qbo_location"), template_values),
            render_program_template(
                output.get("funding_source_template"), template_values
            ),
            render_program_template(
                output.get("service_agreement"), template_values
            ),
            render_program_template(
                output.get("program_initiative_template"), template_values
            ),
        )
        return True

    return False


def _admin_fundraising_mapping_path() -> Path:
    """Return the configured admin and fundraising mapping file."""
    return yaml_mapping_path(ADMIN_FUNDRAISING_MAPPING_FILENAME)


@lru_cache(maxsize=1)
def load_admin_fundraising_rules() -> tuple[dict[str, Any], ...]:
    """Load static admin and fundraising rules from YAML."""
    mapping_path = _admin_fundraising_mapping_path()
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Required admin/fundraising mapping file not found: {mapping_path}"
        )

    with mapping_path.open("r", encoding="utf-8") as stream:
        configuration = yaml.safe_load(stream)

    if not isinstance(configuration, dict):
        raise ValueError(
            f"Admin/fundraising mapping must contain a YAML object: {mapping_path}"
        )

    rules = configuration.get("rules")
    if not isinstance(rules, list) or not rules:
        raise ValueError(
            f"Admin/fundraising mapping must contain a non-empty 'rules' list: "
            f"{mapping_path}"
        )

    for index, rule in enumerate(rules, start=1):
        if not isinstance(rule, dict):
            raise ValueError(f"Admin/fundraising rule {index} must be a YAML object.")
        if not isinstance(rule.get("match"), dict):
            raise ValueError(
                f"Admin/fundraising rule {index} is missing its match block."
            )
        if not isinstance(rule.get("output"), dict):
            raise ValueError(
                f"Admin/fundraising rule {index} is missing its output block."
            )

    return tuple(rules)


def admin_fundraising_rule_matches(
    rule: dict[str, Any],
    derived_class: str,
) -> bool:
    """Return True when Derived_Class satisfies a configured static rule."""
    match = rule["match"]
    match_type = text(match.get("type"))

    if match_type == "exact":
        return derived_class == text(match.get("value"))
    if match_type == "one_of":
        values = match.get("values", [])
        return isinstance(values, list) and derived_class in values

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(
        f"Unsupported admin/fundraising match type in rule {rule_id}: "
        f"{match_type}"
    )


def render_admin_fundraising_value(value: Any) -> str:
    """Resolve shared default tokens used by static mapping outputs."""
    rendered = text(value)
    rendered = rendered.replace("__DEFAULT_CLASS__", DEFAULT_CLASS)
    rendered = rendered.replace("__DEFAULT_LOCATION__", DEFAULT_LOCATION)
    rendered = rendered.replace("__DEFAULT_FUNDING__", DEFAULT_FUNDING)
    return rendered


def handle_admin_and_fundraising_mapping(
    ws: Worksheet,
    row: int,
    derived_class: str,
) -> bool:
    """Apply the first matching rule from qbo_admin_fundraising_mapping.yaml."""
    for rule in load_admin_fundraising_rules():
        if not admin_fundraising_rule_matches(rule, derived_class):
            continue

        output = dict(rule["output"])
        source_overrides = rule.get("source_overrides", {})
        if isinstance(source_overrides, dict):
            override = source_overrides.get(derived_class)
            if isinstance(override, dict):
                output.update(override)

        set_qbo(
            ws,
            row,
            render_admin_fundraising_value(output.get("project")),
            render_admin_fundraising_value(output.get("qbo_class")),
            render_admin_fundraising_value(output.get("location")),
            render_admin_fundraising_value(output.get("funding_source")),
            render_admin_fundraising_value(output.get("service_agreement")),
            render_admin_fundraising_value(output.get("program_initiative")),
        )
        return True

    return False

def _event_mapping_path() -> Path:
    """Return the configured event mapping file."""
    return yaml_mapping_path(EVENT_MAPPING_FILENAME)


@lru_cache(maxsize=1)
def load_event_rules() -> tuple[dict[str, Any], ...]:
    """Load and validate event business rules from qbo_event_mapping.yaml."""
    mapping_path = _event_mapping_path()
    if not mapping_path.exists():
        raise FileNotFoundError(
            f"Required event mapping file not found: {mapping_path}"
        )

    with mapping_path.open("r", encoding="utf-8") as stream:
        configuration = yaml.safe_load(stream)

    if not isinstance(configuration, dict):
        raise ValueError(
            f"Event mapping must contain a YAML object: {mapping_path}"
        )

    events = configuration.get("events")
    if not isinstance(events, list) or not events:
        raise ValueError(
            f"Event mapping must contain a non-empty 'events' list: {mapping_path}"
        )

    for index, rule in enumerate(events, start=1):
        if not isinstance(rule, dict):
            raise ValueError(f"Event rule {index} must be a YAML object.")
        if not isinstance(rule.get("match"), dict):
            raise ValueError(f"Event rule {index} is missing its match block.")
        if not isinstance(rule.get("year"), dict):
            raise ValueError(f"Event rule {index} is missing its year block.")
        if not isinstance(rule.get("output"), dict):
            raise ValueError(f"Event rule {index} is missing its output block.")

    return tuple(events)


def event_rule_matches(rule: dict[str, Any], derived_class: str) -> bool:
    """Return True when a Derived_Class satisfies an event rule."""
    if derived_class in rule.get("exclude_exact", []):
        return False

    match = rule["match"]
    match_type = match.get("type")

    if match_type == "exact":
        return derived_class == text(match.get("value"))
    if match_type == "prefix":
        return starts(derived_class, text(match.get("value")))
    if match_type == "one_of":
        values = match.get("values", [])
        return isinstance(values, list) and derived_class in values

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(f"Unsupported event match type in rule {rule_id}: {match_type}")


def event_rule_year(rule: dict[str, Any], derived_class: str, transaction_date: Any) -> int:
    """Resolve the event year using the rule's fixed, range, or source-text policy."""
    year_rule = rule["year"]
    mode = year_rule.get("mode")

    if mode == "fixed":
        return int(year_rule["value"])

    if mode == "ranges":
        parsed_date = as_date(transaction_date)
        if parsed_date is not None:
            for date_range in year_rule.get("ranges", []):
                range_start = as_date(date_range.get("start"))
                range_end = as_date(date_range.get("end"))
                if range_start is not None and parsed_date < range_start:
                    continue
                if range_end is not None and parsed_date > range_end:
                    continue
                return int(date_range["year"])
        return int(year_rule["fallback"])

    if mode == "source_text":
        allowed_years = [int(value) for value in year_rule.get("allowed_years", [])]
        for year in allowed_years:
            if str(year) in derived_class:
                return year
        return int(year_rule["fallback"])

    rule_id = text(rule.get("id")) or "(unnamed)"
    raise ValueError(f"Unsupported event year mode in rule {rule_id}: {mode}")


def handle_event_mapping(
    ws: Worksheet,
    row: int,
    derived_class: str,
    transaction_date: Any,
) -> bool:
    """Apply the first matching event rule from qbo_event_mapping.yaml."""
    for rule in load_event_rules():
        if not event_rule_matches(rule, derived_class):
            continue

        year = event_rule_year(rule, derived_class, transaction_date)
        output = dict(rule["output"])

        source_overrides = rule.get("source_overrides", {})
        if isinstance(source_overrides, dict):
            override = source_overrides.get(derived_class, {})
            if isinstance(override, dict):
                output.update(override)

        project_template = text(output.get("project_template"))
        if project_template == "":
            rule_id = text(rule.get("id")) or "(unnamed)"
            raise ValueError(f"Event rule {rule_id} has no project_template.")

        project = project_template.format(year=year)
        funding_suffix = text(output.get("funding_suffix")) or DEFAULT_FUNDING
        funding_source = f"{project} | {funding_suffix}"

        set_qbo(
            ws,
            row,
            project,
            text(output.get("qbo_class")) or DEFAULT_CLASS,
            text(output.get("location")) or "Event",
            funding_source,
            text(output.get("service_agreement")),
            text(output.get("program_initiative")),
        )
        return True

    return False


def _organization_registry_path() -> Path:
    """Return the configured organization registry file."""
    return yaml_mapping_path(ORGANIZATION_REGISTRY_FILENAME)


@lru_cache(maxsize=1)
def load_organization_registry() -> dict[str, Any]:
    """Load and validate organization_registry.yaml."""
    registry_path = _organization_registry_path()

    with registry_path.open("r", encoding="utf-8") as stream:
        configuration = yaml.safe_load(stream) or {}

    if not isinstance(configuration, dict):
        raise ValueError(
            f"Organization registry must contain a YAML object: {registry_path}"
        )

    if configuration.get("schema") != "organization_registry":
        raise ValueError(
            f"YAML file is not an organization registry: {registry_path}"
        )

    defaults = configuration.get("defaults", {})
    if defaults is None:
        defaults = {}
    if not isinstance(defaults, dict):
        raise ValueError(
            f"Organization registry defaults must be a YAML object: {registry_path}"
        )

    organizations = configuration.get("organizations")
    if not isinstance(organizations, dict) or not organizations:
        raise ValueError(
            f"Organization registry must contain a non-empty 'organizations' "
            f"object: {registry_path}"
        )

    for stable_id, organization in organizations.items():
        if not isinstance(stable_id, str) or not stable_id.strip():
            raise ValueError("Organization registry stable IDs must be nonblank text.")
        if not isinstance(organization, dict):
            raise ValueError(
                f"Organization {stable_id!r} must be a YAML object."
            )

        name = organization.get("name")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(
                f"Organization {stable_id!r} requires a nonblank name."
            )

        approved = organization.get("approved")
        if not isinstance(approved, bool):
            raise ValueError(
                f"Organization {stable_id!r} requires a boolean approved value."
            )

        matches = organization.get("matches")
        if not isinstance(matches, list) or not matches:
            raise ValueError(
                f"Organization {stable_id!r} requires a non-empty matches list."
            )

        for index, match in enumerate(matches, start=1):
            if not isinstance(match, dict):
                raise ValueError(
                    f"Organization {stable_id!r} match {index} must be an object."
                )
            match_type = text(match.get("type"))
            if match_type not in {"exact", "prefix"}:
                raise ValueError(
                    f"Organization {stable_id!r} match {index} has unsupported "
                    f"type: {match_type!r}."
                )
            if text(match.get("value")) == "":
                raise ValueError(
                    f"Organization {stable_id!r} match {index} requires a value."
                )

        funding = organization.get("funding", {})
        if not isinstance(funding, dict):
            raise ValueError(
                f"Organization {stable_id!r} funding must be an object."
            )
        for field_name in ("exact", "date_based"):
            field_value = funding.get(field_name, {})
            if not isinstance(field_value, dict):
                raise ValueError(
                    f"Organization {stable_id!r} funding.{field_name} "
                    "must be an object."
                )

        for field_name in ("service_agreements", "program_initiatives"):
            field_value = organization.get(field_name, {})
            if not isinstance(field_value, dict):
                raise ValueError(
                    f"Organization {stable_id!r} {field_name} must be an object."
                )

    return configuration


def organization_rule_matches(
    match: dict[str, Any],
    derived_class: str,
) -> bool:
    """Return True when Derived_Class satisfies one registry match rule."""
    match_type = text(match.get("type"))
    expected = text(match.get("value"))

    if match_type == "exact":
        return derived_class == expected
    if match_type == "prefix":
        return starts(derived_class, expected)

    raise ValueError(f"Unsupported organization match type: {match_type}")


def match_organization(
    derived_class: str,
) -> Optional[tuple[str, dict[str, Any]]]:
    """Return the stable ID and registry record for the matching organization."""
    organizations = load_organization_registry()["organizations"]

    for stable_id, organization in organizations.items():
        matches = organization["matches"]
        if any(
            organization_rule_matches(match, derived_class)
            for match in matches
        ):
            return str(stable_id), organization

    return None


def resolve_organization_date_based_funding(
    organization: dict[str, Any],
    derived_class: str,
    transaction_date: Any,
) -> Optional[str]:
    """Resolve a date-based funding override from one organization record."""
    funding = organization.get("funding", {})
    mappings = funding.get("date_based", {})

    rule = mappings.get(derived_class)
    if not isinstance(rule, dict):
        return None

    parsed_date = as_date(transaction_date)
    ranges = rule.get("ranges", [])
    if not isinstance(ranges, list):
        raise ValueError(
            f"Date-based funding ranges must be a list for {derived_class}."
        )

    if parsed_date is not None:
        for date_range in ranges:
            if not isinstance(date_range, dict):
                raise ValueError(
                    f"Each date-based funding range must be an object for "
                    f"{derived_class}."
                )
            start_date = as_date(date_range.get("start"))
            if start_date is not None and parsed_date >= start_date:
                return text(date_range.get("funding"))

    return text(rule.get("fallback"))


def default_organization_funding(
    derived_class: str,
    organization: dict[str, Any],
) -> str:
    """Resolve funding from one organization registry record."""
    registry = load_organization_registry()
    defaults = registry.get("defaults", {})
    funding = organization.get("funding", {})

    exact = funding.get("exact", {})
    if derived_class in exact:
        return text(exact[derived_class])

    suffixes = defaults.get("general_suffixes", [])
    if not isinstance(suffixes, list):
        raise ValueError(
            "Organization registry defaults.general_suffixes must be a list."
        )
    if any(derived_class.endswith(text(suffix)) for suffix in suffixes):
        return DEFAULT_FUNDING

    organization_name = text(organization.get("name"))
    if organization_name and starts(derived_class, organization_name):
        remainder = derived_class[len(organization_name):].strip()
        if remainder.startswith("-"):
            remainder = remainder[1:].strip()

        prefixes = defaults.get("known_funding_prefixes", [])
        if not isinstance(prefixes, list):
            raise ValueError(
                "Organization registry defaults.known_funding_prefixes "
                "must be a list."
            )
        if any(remainder.startswith(text(prefix)) for prefix in prefixes):
            return remainder

    configured = text(funding.get("default"))
    if configured == "":
        configured = text(defaults.get("funding_source"))
    if configured in {"", "__DEFAULT_FUNDING__"}:
        return DEFAULT_FUNDING
    return configured


def organization_funding_source(
    organization: str,
    funding_source: str,
) -> str:
    """Return the standard organization-qualified QBO funding source."""
    funding_source = text(funding_source) or DEFAULT_FUNDING
    return f"{organization} | {funding_source}"


def handle_organizations(
    ws: Worksheet,
    row: int,
    derived_class: str,
    transaction_date: Any,
    distribution_account: str,
) -> bool:
    """Apply mappings from organization_registry.yaml."""
    match = match_organization(derived_class)
    if match is None:
        return False

    _, organization_record = match
    organization_name = text(organization_record.get("name"))

    sdg_mapping = resolve_sdg_mapping(
        ws,
        row,
        derived_class,
        distribution_account,
        transaction_date,
    )

    services = organization_record.get("service_agreements", {})
    initiatives = organization_record.get("program_initiatives", {})

    service = text(services.get(derived_class))
    initiative = ""

    if sdg_mapping is not None:
        funding = organization_funding_source(
            organization_name,
            sdg_mapping["funding_base"],
        )
        initiative = sdg_mapping["program_initiative"]
    else:
        initiative = text(initiatives.get(derived_class))
        dated_funding = resolve_organization_date_based_funding(
            organization_record,
            derived_class,
            transaction_date,
        )
        base_funding = (
            dated_funding
            if dated_funding is not None
            else default_organization_funding(
                derived_class,
                organization_record,
            )
        )
        funding = organization_funding_source(
            organization_name,
            base_funding,
        )

    set_qbo(
        ws,
        row,
        organization_name,
        DEFAULT_CLASS,
        DEFAULT_LOCATION,
        funding,
        service,
        initiative,
    )
    return True

def process_row(ws: Worksheet, row: int) -> str:
    #data_row = ws[f"{COL['Data_Row_1']}{row}"].value

    derived_class = text(ws[f"{COL['Derived_Class']}{row}"].value)
    
    # print(f"Processing: {derived_class}" )

    #if not is_number(data_row) or derived_class == "":
    if derived_class == "":
        return "skipped"

    transaction_date = ws[f"{COL['Transaction_Date']}{row}"].value
    distribution_account = text(ws[f"{COL['Distribution_Account']}{row}"].value)

    # The legacy mapping rules did not explicitly populate QBO_Account.
    # Carry the source distribution account into the new QBO account field.
    ws[f"{COL['QBO_Account']}{row}"] = clean_deleted_marker(distribution_account)

    preserve_legacy_detail(ws, row)

    if handle_program_mapping(ws, row, derived_class, transaction_date):
        return "mapped"
    if handle_standalone_sdg_mapping(
        ws,
        row,
        derived_class,
        distribution_account,
        transaction_date,
    ):
        return "mapped"
    if handle_admin_and_fundraising_mapping(ws, row, derived_class):
        return "mapped"
    if handle_event_mapping(ws, row, derived_class, transaction_date):
        return "mapped"
    if handle_organizations(
        ws, row, derived_class, transaction_date, distribution_account
    ):
        return "mapped"

    ws[f"{COL['Review_Flag']}{row}"] = f"REVIEW: {derived_class}"
    print(f"UNMAPPED: {derived_class}")
    return "unmapped"


def dimension_upload(
    workbook_path: Path,
    yaml_directory: Path,
) -> dict[str, int]:
    """Populate QBO dimensions using the organization registry and other mappings."""
    configure_yaml_directory(yaml_directory)
    clear_mapping_caches()

    workbook_path = Path(workbook_path)

    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Dimensioner input must be an .xlsx workbook.")
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if not workbook_path.is_file():
        raise ValueError(f"Dimensioner input is not a file: {workbook_path}")

    workbook = load_workbook(workbook_path)
    try:
        if workbook.sheetnames != ["Upload"]:
            raise ValueError(
                'Dimensioner input must contain exactly one worksheet named "Upload"; '
                f"found: {', '.join(workbook.sheetnames) or '(none)'}"
            )

        worksheet = workbook["Upload"]

        required_headers = {name: column for name, column in COL.items()}
        missing_headers = [
            name
            for name, column in required_headers.items()
            if text(worksheet[f"{column}1"].value) != name
        ]
        if missing_headers:
            raise ValueError(
                "Upload worksheet is missing or has misplaced required columns: "
                + ", ".join(missing_headers)
            )

        counts = {"mapped": 0, "unmapped": 0, "skipped": 0}
        for row in range(2, worksheet.max_row + 1):
            result = process_row(worksheet, row)
            counts[result] += 1

        workbook.save(workbook_path)
        return counts
    finally:
        workbook.close()


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(
            "ERROR: Expected two arguments: Upload workbook path and YAML directory.",
            file=sys.stderr,
        )
        return 2

    workbook_path = Path(argv[0]).expanduser().resolve()
    yaml_directory = Path(argv[1]).expanduser().resolve()
    try:
        counts = dimension_upload(workbook_path, yaml_directory)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Updated: {workbook_path}")
    print(f"Rows mapped: {counts['mapped']}")
    print(f"Rows unmapped: {counts['unmapped']}")
    print(f"Rows skipped: {counts['skipped']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
